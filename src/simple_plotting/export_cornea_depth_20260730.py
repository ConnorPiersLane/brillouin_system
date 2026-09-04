"""Excel export of every point behind `plot_cornea_depth_20260730.py`.

One row per scan (each scan is a single frame), including the ones the strict
gate threw out, so the rejection is auditable rather than invisible. Fitting
model, D and the gate all come from the plotting module -- change them there and
re-run this, so the figure and the workbook can never disagree.

Sheets:
    Data     every scan, kept or dropped, with the reason
    Summary  n / mean / sd / sem per file, as live formulas over Data
    Notes    what each column is and how the numbers were produced

Usage:
    PYTHONPATH=src python src/simple_plotting/export_cornea_depth_20260730.py [out_dir]
"""
import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from plot_cornea_depth_20260730 import (  # noqa: E402
    DATA, FITTING_MODEL, MAX_LR_DISAGREE, MAX_PLANE_GAP_UM, N_CORNEA, N_SAMPLE,
    PEOPLE, PLAUSIBLE, REFERENCE_MODEL, SESSION_D_MM,
    fit_scan, load_scans, planes, session_fitter,
)

OUT = Path(sys.argv[1] if len(sys.argv) > 1 else ".")

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="DDE5EE")
KEPT_FILL = PatternFill("solid", fgColor="E8F3E8")
THIN = Side(style="thin", color="BFBFBF")

COLUMNS = [
    ("Subject", 10, None),
    ("File", 15, None),
    ("Scan i", 7, "0"),
    ("Commanded offset (um)", 12, "0"),
    ("Forward plane z (um)", 13, "0.0"),
    ("Backward plane z (um)", 13, "0.0"),
    ("Plane gap fwd-bwd (um)", 13, "0.0"),
    ("Lens z (um)", 12, "0.0"),
    ("Depth, lens travel (um)", 13, "0.0"),
    ("Depth in tissue (um)", 13, "0.0"),
    ("Brillouin shift (GHz)", 13, "0.0000"),
    ("Per-frame sigma (MHz)", 13, "0.00"),
    ("Left peak (GHz)", 12, "0.0000"),
    ("Right peak (GHz)", 12, "0.0000"),
    ("Left-Right (MHz)", 12, "0.0"),
    ("Photons, left", 11, "0"),
    ("Photons, right", 11, "0"),
    ("Kept", 7, None),
    ("Reason", 30, None),
    # The two columns the Summary averages. The shift column above is filled for
    # any scan that FITTED, including ones the plane gate rejected -- averaging
    # that column would silently include points the figure never plotted.
    ("Shift, kept only (GHz)", 13, "0.0000"),
    ("Sigma, kept only (MHz)", 13, "0.00"),
]


def rows_for(subject: str, filename: str, sf):
    """Every scan in one file, with the same gate the figure applies."""
    out = []
    for scan in load_scans(DATA / filename):
        fz, bz = planes(scan)
        gap = (fz - bz) if (fz is not None and bz is not None) else None
        z_lens = float(scan.measurements[0].lens_zaber_position)
        z_off = (float(scan.reflection_result_forwards.z_offset_um)
                 if (scan.reflection_result_forwards is not None
                     and scan.reflection_result_forwards.z_offset_um is not None) else None)

        s = fit_scan(scan, sf)[0]
        ok = s.fitted_spectrum.is_success
        a = s.analyzed_shifts

        def num(v):
            return None if v is None or not np.isfinite(v) else float(v)

        shift = num(a.freq_shift_peak_distance_ghz) if ok else None
        left = num(a.freq_shift_left_peak_ghz) if ok else None
        right = num(a.freq_shift_right_peak_ghz) if ok else None
        sigma = num(s.theoretical_precisions.distance_total_mhz) if ok else None
        ph_l = num(getattr(s.photons, "left_peak_photons", None)) if ok else None
        ph_r = num(getattr(s.photons, "right_peak_photons", None)) if ok else None
        lr = (left - right) * 1000.0 if (left is not None and right is not None) else None

        pass_plane = gap is not None and abs(gap) <= MAX_PLANE_GAP_UM
        pass_fit = (shift is not None and left is not None and right is not None
                    and abs(left - right) < MAX_LR_DISAGREE
                    and PLAUSIBLE[0] < shift < PLAUSIBLE[1])
        depth = (z_lens - 0.5 * (fz + bz)) if gap is not None else None

        if gap is None:
            reason = "no backward plane"
        elif not pass_plane:
            reason = f"planes disagree {gap:+.0f} um"
        elif shift is None:
            reason = "fit failed"
        elif not (PLAUSIBLE[0] < shift < PLAUSIBLE[1]):
            reason = "shift outside 4-8 GHz (elastic reflection)"
        elif abs(left - right) >= MAX_LR_DISAGREE:
            reason = f"left-right disagree {lr:+.0f} MHz"
        else:
            reason = "kept"

        out.append([
            subject, filename.replace(".h5", ""), scan.i, z_off, fz, bz, gap, z_lens,
            depth, (depth * N_CORNEA if depth is not None else None),
            shift if pass_fit else None, sigma if pass_fit else None,
            left, right, lr, ph_l, ph_r,
            "yes" if (pass_plane and pass_fit) else "no", reason,
            shift if (pass_plane and pass_fit) else None,
            sigma if (pass_plane and pass_fit) else None,
        ])
    return out


def write_data(ws, rows):
    ws.append([c[0] for c in COLUMNS])
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True, size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[1].height = 32

    for r in rows:
        ws.append(r)
    for i, (_, width, fmt) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
        for cell in ws[get_column_letter(i)][1:]:
            cell.font = Font(name=FONT, size=10)
            if fmt:
                cell.number_format = fmt
    # Shade the rows that actually made it into the figure.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[17].value == "yes":
            for cell in row:
                cell.fill = KEPT_FILL
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"


def write_summary(ws, blocks):
    """Per-file statistics as formulas over Data, so edits to Data propagate.

    `blocks` is [(file, first_row, last_row)] and the rows of each file are
    contiguous on Data, so every statistic is a plain range formula -- no array
    formulas, which openpyxl cannot mark as CSE and Excel would then mis-evaluate.
    AVERAGE / STDEV / COUNT skip the blanks left by rejected scans, which is
    exactly the wanted behaviour.
    """
    ws.append(["File", "Scans", "Pass plane gate", "Kept (plotted)",
               "Mean shift (GHz)", "sd (MHz)", "sem (MHz)", "Mean per-frame sigma (MHz)"])
    for cell in ws[1]:
        cell.font = Font(name=FONT, bold=True, size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
        cell.border = Border(bottom=THIN)
    ws.row_dimensions[1].height = 32

    g = int(MAX_PLANE_GAP_UM)
    for i, (fn, r0, r1) in enumerate(blocks, start=2):
        ws.append([
            fn.replace(".h5", ""),
            f'=ROWS(Data!$A{r0}:$A{r1})',
            f'=COUNTIFS(Data!$G{r0}:$G{r1},">=-{g}",Data!$G{r0}:$G{r1},"<={g}")',
            f'=COUNTIF(Data!$R{r0}:$R{r1},"yes")',
            f'=AVERAGE(Data!$T{r0}:$T{r1})',
            f'=IF($D{i}>1,STDEV(Data!$T{r0}:$T{r1})*1000,"")',
            f'=IF($D{i}>1,$F{i}/SQRT($D{i}),"")',
            f'=AVERAGE(Data!$U{r0}:$U{r1})',
        ])

    for col, width, fmt in [("A", 16, None), ("B", 8, "0"), ("C", 13, "0"), ("D", 13, "0"),
                            ("E", 14, "0.0000"), ("F", 10, "0.0"), ("G", 10, "0.0"),
                            ("H", 14, "0.00")]:
        ws.column_dimensions[col].width = width
        for cell in ws[col][1:]:
            cell.font = Font(name=FONT, size=10)
            if fmt:
                cell.number_format = fmt


def write_notes(ws, stats):
    lines = [
        ("Cornea Brillouin shift vs. depth - 2026-07-30", True),
        ("", False),
        ("Source files", True),
        (f"{DATA}", False),
        ("selina_50.h5, selina_100um.h5, yuxuan_50um.h5, yuxuan_100um.h5", False),
        ("", False),
        ("How each row was produced", True),
        ("Each scan in these files is ONE camera frame taken at a commanded offset past the "
         "surface found by the reflection finder on the way in (forward). The finder runs "
         "again on the way out (backward), giving an independent estimate of the same "
         "surface.", False),
        (f"Fitting model: {FITTING_MODEL}, calibration fitted with '{REFERENCE_MODEL}' "
         f"(the two must be the same lineshape family; mixing them moves every shift by "
         f"~3 MHz).", False),
        (f"NA correction: D (na_beam_diameter_mm) = {SESSION_D_MM} mm, solved on this "
         f"session's own two-NA water bracket water_na042_na014.h5, where na042 and na014 "
         f"both land on f180 = 5.0704 GHz (residual gap 0.000 MHz, raw gap -14.1 MHz). "
         f"n_sample = {N_SAMPLE} (cornea).", False),
        ("D is common-mode: it slides every point together by ~3.9 MHz per mm and cannot "
         "create or remove a depth trend.", False),
        ("", False),
        ("The strict gate (three conditions, all required)", True),
        (f"1. Both reflection planes found, and |forward - backward| <= {MAX_PLANE_GAP_UM:.0f} um.",
         False),
        (f"2. The fit succeeded and the shift is inside {PLAUSIBLE[0]}-{PLAUSIBLE[1]} GHz. "
         f"Frames failing this landed on the elastic reflection (shift ~ -0.2 GHz).", False),
        (f"3. Left and right peaks agree within {MAX_LR_DISAGREE*1000:.0f} MHz.", False),
        ("", False),
        ("Column meanings that are easy to misread", True),
        ("Depth, lens travel (um) = lens z - (forward + backward)/2. The pair average is the "
         "surface: averaging the two sweep directions cancels the latency bias.", False),
        (f"Depth in tissue (um) = lens travel x n = {N_CORNEA}. Approximate - focus moves "
         f"roughly n times deeper than the stage.", False),
        ("Per-frame sigma (MHz) = theoretical shot-noise precision on the inter-peak "
         "distance (Thompson/CRLB from the fitted widths and photon counts). This is the "
         "error bar drawn on each point in the figure. It is NOT repeatability - each row is "
         "a single frame with no repeat.", False),
        ("sd on the Summary sheet = scatter of the plotted points about their own mean. It "
         "runs 2-5x larger than the per-frame sigma, so these measurements are NOT photon "
         "limited. Quote the sd, never the per-frame sigma, as the uncertainty of a mean.",
         False),
        ("", False),
        ("Values as generated (a cross-check on the Summary formulas)", True),
    ]
    lines += [(text, False) for text in stats]
    lines += [
        ("", False),
        ("Regenerate", True),
        ("PYTHONPATH=src python src/simple_plotting/plot_cornea_depth_20260730.py <out_dir>",
         False),
        ("PYTHONPATH=src python src/simple_plotting/export_cornea_depth_20260730.py <out_dir>",
         False),
    ]
    for text, bold in lines:
        ws.append([text])
        c = ws.cell(row=ws.max_row, column=1)
        c.font = Font(name=FONT, size=10, bold=bold)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 118


if __name__ == "__main__":
    OUT.mkdir(parents=True, exist_ok=True)
    sf = session_fitter()

    all_rows, blocks, stats = [], [], []
    for subject, filenames in PEOPLE.items():
        for fn in filenames:
            rows = rows_for(subject, fn, sf)
            first = len(all_rows) + 2          # +1 for the header, +1 for 1-indexing
            all_rows += rows
            blocks.append((fn, first, len(all_rows) + 1))

            y = np.array([r[10] for r in rows if r[17] == "yes"], float)
            if y.size > 1:
                stats.append(f"{fn.replace('.h5','')}: n = {y.size}, mean = {y.mean():.4f} GHz, "
                             f"sd = {y.std(ddof=1)*1000:.1f} MHz, "
                             f"sem = {y.std(ddof=1)/np.sqrt(y.size)*1000:.1f} MHz")
            elif y.size == 1:
                stats.append(f"{fn.replace('.h5','')}: n = 1, value = {y[0]:.4f} GHz "
                             f"(no sd from a single point)")
            print(f"{fn}: {len(rows)} scans, {y.size} kept")

    wb = Workbook()
    write_data(wb.active, all_rows)
    wb.active.title = "Data"
    write_summary(wb.create_sheet("Summary"), blocks)
    write_notes(wb.create_sheet("Notes"), stats)

    out = OUT / "cornea_depth_20260730.xlsx"
    wb.save(out)
    print(f"-> {out}")
